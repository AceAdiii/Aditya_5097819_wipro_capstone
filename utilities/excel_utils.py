import os
from openpyxl import load_workbook


def get_excel_data():
    base_dir = os.path.dirname(os.path.dirname(__file__))
    file_path = os.path.join(base_dir, "TestData", "bus_test_data.xlsx")

    wb = load_workbook(file_path)
    sheet = wb.active

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)

    return data